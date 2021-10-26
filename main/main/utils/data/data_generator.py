import re
from datetime import datetime

from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Permission, User
from django.http import HttpResponse
from openpyxl import load_workbook

from main.models.cost_model import CostModel
from main.models.customer_model import CustomerModel
from main.models.nhap_kho_model import NhapKhoModel
from main.models.po_model import POModel
from main.models.xuat_kho_model import XuatKhoModel
from main.serializers.bosch_revenue_ser import BoschRevenueSerializer
from main.serializers.cost_ser import CostSerializer
from main.serializers.customer_ser import CustomerSerializer
from main.serializers.nhap_kho_ser import NhapKhoSerializer
from main.serializers.po_ser import POSerializer
from main.serializers.side_revenue_ser import SideRevenueSerializer
from main.serializers.xuat_kho_ser import XuatKhoSerializer
from main.utils.pricing_service import calculate_gia_goc_xuat_kho


def generate_test_data(request):
    loc = "main/utils/data/data_1.xlsx"
    # To open Workbook
    wb = load_workbook(loc, data_only=True)

    if not POModel.objects.all():
        # PO
        sheet_obj = wb.worksheets[0]
        row = 7
        while True:
            pn_13 = sheet_obj.cell(row=row, column=2).value
            if not pn_13:
                break
            validated_data = {
                "po_no": sheet_obj.cell(row=row, column=1).value.strip() or '',
                "pn_13": str(sheet_obj.cell(row=row, column=2).value).strip() or '',
                "pn_10": sheet_obj.cell(row=row, column=3).value.strip() or '',
                "bosch_no": str(sheet_obj.cell(row=row, column=4).value).strip() or '',
                "z_exel_no": sheet_obj.cell(row=row, column=5).value.strip() if sheet_obj.cell(row=row,
                                                                                               column=5).value else '',
                "stamping": sheet_obj.cell(row=row, column=6).value.strip() if sheet_obj.cell(row=row,
                                                                                              column=6).value else '',
                "country": sheet_obj.cell(row=row, column=7).value.strip() if sheet_obj.cell(row=row,
                                                                                             column=7).value else '',
                "english_des": sheet_obj.cell(row=row, column=8).value.strip() if sheet_obj.cell(row=row,
                                                                                                 column=8).value else '',
                "import_des": sheet_obj.cell(row=row, column=9).value.strip() if sheet_obj.cell(row=row,
                                                                                                column=9).value else '',
                "app_des": sheet_obj.cell(row=row, column=10).value.strip() if sheet_obj.cell(row=row,
                                                                                              column=10).value else '',
                "quantity": sheet_obj.cell(row=row, column=11).value or 0,
                "dap_price": round(sheet_obj.cell(row=row, column=12).value or 0, 3),
                "extension_price": round(sheet_obj.cell(row=row, column=13).value or 0, 3),
                "tax": round(sheet_obj.cell(row=row, column=14).value or 0, 3),
                "gia_von": round(sheet_obj.cell(row=row, column=15).value or 0, 3),
                "gia_si": round(sheet_obj.cell(row=row, column=16).value or 0, 3),
                "gia_le": round(sheet_obj.cell(row=row, column=17).value or 0, 3),
                "lead_time": sheet_obj.cell(row=row, column=18).value.strip() if sheet_obj.cell(row=row,
                                                                                                column=18).value else '',
                "customer": sheet_obj.cell(row=row, column=19).value.strip() if sheet_obj.cell(row=row,
                                                                                               column=19).value else '',
                "remarks": sheet_obj.cell(row=row, column=20).value.strip() if sheet_obj.cell(row=row,
                                                                                              column=20).value else '',
            }
            if not validated_data["gia_le"]:
                validated_data["gia_le"] = validated_data["gia_si"] * 1.1
            POSerializer().create(validated_data)
            row += 1

    if not NhapKhoModel.objects.all():
        # Nhap kho
        sheet_obj = wb.worksheets[1]
        row = 4
        while True:
            pn_13 = sheet_obj.cell(row=row, column=4).value
            if not pn_13:
                break
            input_date = None
            cell_value = sheet_obj.cell(row=row, column=1).value
            if type(cell_value) == datetime:
                input_date = cell_value.strftime("%Y-%m-%d")
            validated_data = {
                "input_date": input_date,
                "license_no": sheet_obj.cell(row=row, column=2).value if sheet_obj.cell(row=row,
                                                                                        column=2).value else '',
                "provider": sheet_obj.cell(row=row, column=3).value.strip() or '',
                "pn_13": sheet_obj.cell(row=row, column=4).value.strip() or '',
                "pn_10": sheet_obj.cell(row=row, column=5).value.strip() or '',
                "bosch_no": str(sheet_obj.cell(row=row, column=6).value).strip() if sheet_obj.cell(row=row,
                                                                                                   column=6).value else '',
                "z_exel_no": sheet_obj.cell(row=row, column=7).value.strip() if sheet_obj.cell(row=row,
                                                                                               column=7).value else '',
                "stamping": sheet_obj.cell(row=row, column=8).value.strip() if sheet_obj.cell(row=row,
                                                                                              column=8).value else '',
                "gia_von": sheet_obj.cell(row=row, column=11).value or 0,
                "gia_si": sheet_obj.cell(row=row, column=12).value or 0,
                "gia_le": sheet_obj.cell(row=row, column=13).value or 0,
                "quantity": sheet_obj.cell(row=row, column=14).value or 0,
                "dap_price": round(float(sheet_obj.cell(row=row, column=15).value or 0), 3),
                "extension_price": round(sheet_obj.cell(row=row, column=16).value or 0, 3),
                "ratio": sheet_obj.cell(row=row, column=17).value or 0,
                "tax": float(sheet_obj.cell(row=row, column=18).value or 0),
                "vat_percentage": int(re.search(r'\d+', sheet_obj.cell(row=row, column=19).value or 0).group()),
                "po_no": sheet_obj.cell(row=row, column=20).value.strip() if sheet_obj.cell(row=row,
                                                                                            column=20).value else '',
            }
            validated_data["gia_goc"] = round(
                validated_data["dap_price"] * validated_data["ratio"] * validated_data["tax"], 3)
            validated_data["tong_gia_goc"] = round(validated_data["gia_goc"] * validated_data["quantity"], 3)
            if not validated_data["gia_le"]:
                validated_data["gia_le"] = validated_data["gia_si"] * 1.1

            recent_po = POModel.objects.filter(pn_13=validated_data["pn_13"]).last()
            if recent_po:
                validated_data["english_des"] = recent_po.english_des
                validated_data["app_des"] = recent_po.app_des
                validated_data["import_des"] = recent_po.import_des
            NhapKhoSerializer().create(validated_data)
            row += 1

    if not XuatKhoModel.objects.all():
        # Xuat kho
        sheet_obj = wb.worksheets[2]
        row = 8
        while True:
            pn_13 = sheet_obj.cell(row=row, column=3).value.strip() if sheet_obj.cell(row=row, column=3).value else ""
            if not pn_13:
                break
            input_date = sheet_obj.cell(row=row, column=2).value
            if input_date:
                input_date = input_date.strftime("%Y-%m-%d")

            if pn_13 == "0445020126391":
                pass
            validated_data = {
                "input_date": input_date,
                "pn_13": pn_13,
                "quantity": sheet_obj.cell(row=row, column=6).value if sheet_obj.cell(row=row, column=6).value else '',
            }
            last_nhap_kho = NhapKhoModel.objects.filter(pn_13=validated_data["pn_13"]).last()
            if last_nhap_kho:
                validated_data["gia_goc"] = calculate_gia_goc_xuat_kho(pn_13=pn_13, calculate_date=input_date)
                validated_data["gia_si"] = last_nhap_kho.gia_si
                validated_data["gia_le"] = last_nhap_kho.gia_le
                validated_data["english_des"] = last_nhap_kho.english_des
                validated_data["import_des"] = last_nhap_kho.import_des
                validated_data["app_des"] = last_nhap_kho.app_des
                validated_data["tien_goc"] = round(validated_data["gia_goc"] * validated_data["quantity"], 3)
                validated_data["customer"] = ""

            XuatKhoSerializer().create(validated_data)
            row += 1

    # Chi Phi
    loc = "main/utils/data/chi_phi.xlsx"
    # To open Workbook
    wb = load_workbook(loc, data_only=True)

    if not CostModel.objects.all():
        # CPBANHANG
        sheet_obj = wb.worksheets[0]
        row = 8
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            major = sheet_obj.cell(row=row, column=3).value
            client_code = sheet_obj.cell(row=row, column=4).value
            content = sheet_obj.cell(row=row, column=5).value
            value = sheet_obj.cell(row=row, column=6).value
            cost_type = 'CPBANHANG'

            validated_data = {
                "date": date,
                "major": major,
                "client_code": client_code,
                "content": content,
                "value": value,
                "cost_type": cost_type,
            }
            CostSerializer().create(validated_data)
            row += 1

        sheet_obj = wb.worksheets[1]
        row = 7
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            major = sheet_obj.cell(row=row, column=3).value
            client_code = sheet_obj.cell(row=row, column=4).value
            content = sheet_obj.cell(row=row, column=5).value
            value = sheet_obj.cell(row=row, column=6).value
            cost_type = 'CPQUANLY'

            validated_data = {
                "date": date,
                "major": major,
                "client_code": client_code,
                "content": content,
                "value": value,
                "cost_type": cost_type,
            }
            CostSerializer().create(validated_data)
            row += 1

        sheet_obj = wb.worksheets[2]
        row = 7
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            major = sheet_obj.cell(row=row, column=3).value
            client_code = sheet_obj.cell(row=row, column=4).value
            content = sheet_obj.cell(row=row, column=5).value
            value = sheet_obj.cell(row=row, column=6).value
            cost_type = 'CPKHAC'

            validated_data = {
                "date": date,
                "major": major,
                "client_code": client_code,
                "content": content,
                "value": value,
                "cost_type": cost_type,
            }
            CostSerializer().create(validated_data)
            row += 1

        sheet_obj = wb.worksheets[3]
        row = 8
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            major = sheet_obj.cell(row=row, column=3).value
            client_code = sheet_obj.cell(row=row, column=4).value
            content = sheet_obj.cell(row=row, column=5).value
            value = sheet_obj.cell(row=row, column=6).value
            cost_type = 'THUCTECHI'

            validated_data = {
                "date": date,
                "major": major,
                "client_code": client_code,
                "content": content,
                "value": value,
                "cost_type": cost_type,
            }
            CostSerializer().create(validated_data)
            row += 1

        sheet_obj = wb.worksheets[4]
        row = 7
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            client_code = sheet_obj.cell(row=row, column=3).value
            product_code = sheet_obj.cell(row=row, column=4).value
            quantity = sheet_obj.cell(row=row, column=5).value or 0
            price = sheet_obj.cell(row=row, column=6).value
            total = sheet_obj.cell(row=row, column=7).value

            validated_data = {
                "date": date,
                "product_code": product_code,
                "client_code": client_code,
                "quantity": quantity,
                "price": price,
                "total": total,
            }
            BoschRevenueSerializer().create(validated_data)
            row += 1

        sheet_obj = wb.worksheets[5]
        row = 7
        while True:
            date = sheet_obj.cell(row=row, column=2).value
            if not date:
                break
            client_code = sheet_obj.cell(row=row, column=3).value
            accessory_code = sheet_obj.cell(row=row, column=4).value
            quantity = sheet_obj.cell(row=row, column=5).value or 0
            gia_von = sheet_obj.cell(row=row, column=6).value
            tien_von = sheet_obj.cell(row=row, column=7).value
            gia_ban = sheet_obj.cell(row=row, column=8).value
            tien_ban = sheet_obj.cell(row=row, column=9).value
            content = sheet_obj.cell(row=row, column=10).value
            validated_data = {
                "date": date,
                "client_code": client_code,
                "accessory_code": accessory_code,
                "quantity": quantity,
                "gia_von": gia_von,
                "tien_von": tien_von,
                "gia_ban": gia_ban,
                "tien_ban": tien_ban,
                "content": content,
            }
            SideRevenueSerializer().create(validated_data)
            row += 1
    if not CustomerModel.objects.all():
        customer_data = [{"code": "BOSUATAYNINH", "name": "Cty Bò sữa Tây Ninh"},
                         {"code": "CHIBOMBEC", "name": "Cty TNHH Đồng Chí - Anh Chí"},
                         {"code": "CHUONGDIEN", "name": "Chương thợ điện"},
                         {"code": "CUONGBOMBEC", "name": "Cương _Đà Nẵng"},
                         {"code": "DATBOMBEC", "name": "Đạt_QL13"},
                         {"code": "DONGBOMBEC", "name": "Cty TNHH Đồng Chí - Anh Đồng"},
                         {"code": "HIEPHP", "name": "Hiệp- Hải Phòng"},
                         {"code": "HOAVINH", "name": "Hòa vinh _ Của hàng"},
                         {"code": "HUONGHD", "name": "Hương_Garaga_Hải_Dương"},
                         {"code": "KIMBOMBEC", "name": "Kim_Đà_Lạt"},
                         {"code": "LETLT", "name": "Lễ Cty TLT"},
                         {"code": "LYBIENHOA", "name": "Ly_Biên_Hoà"},
                         {"code": "NHIENBOMBEC", "name": "Anh Nhiên"},
                         {"code": "SONHUE", "name": "Sơn huế _ Garage"},
                         {"code": "TAMBOMBEC", "name": "Cty TNHH Phong Tâm"},
                         {"code": "THAITLT", "name": "Thái Cty TLT"},
                         {"code": "THANHTRIEU", "name": "Thanh Triều _ Của Hàng"},
                         {"code": "TLT", "name": "TLT"},
                         {"code": "TUANBOMBEC", "name": "Tuấn_Hai_Đầy"},
                         {"code": "TUANTLT", "name": "Tuấn Cty TLT"},
                         {"code": "TUNGBOMBEC", "name": "Tùng_Hải_Phòng"},
                         {"code": "VINABOSCH", "name": "Vinabosch"},
                         {"code": "TONGLONGAN", "name": "Tòng garage Long An"},
                         {"code": "BOSUATHANHHOA", "name": "Cty Bò sữa Thanh Hoá"}]
        for validated_data in customer_data:
            CustomerSerializer().create(validated_data)

    # Create group-permission
    # Create admin
    if not User.objects.filter(username='admin'):

        u = User.objects.create(username='admin', password=make_password('adminpass'), is_superuser=True)
        permissions = Permission.objects.all()
        for p in permissions:
            u.user_permissions.add(p)
            u.save()
    return HttpResponse("Done")
