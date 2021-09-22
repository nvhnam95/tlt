from io import StringIO
from zipfile import ZipFile
import tempfile

from django.core.files.temp import NamedTemporaryFile
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import pandas
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook

from pyexcel_xlsx import save_data

keys = []
wb = Workbook()
ws = wb.active


def json_to_excel(request):
    json_data = [{"name": "Nam"}]
    for i in range(len(json_data)):
        sub_obj = json_data[i]
        if i == 0:
            keys = list(sub_obj.keys())
            for k in range(len(keys)):
                # row or column index start from 1
                ws.cell(row=(i + 1), column=(k + 1), value=keys[k])
        for j in range(len(keys)):
            ws.cell(row=(i + 2), column=(j + 1), value=sub_obj[keys[j]])
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)

    return HttpResponse(virtual_workbook)
